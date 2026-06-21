#!/usr/bin/env python3
"""Import recordings DATA from Graham McDonald's Excel spreadsheet into recordings.html."""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve.parent.parent
RECORDINGS = ROOT / "recordings.html"
SEP = re.compile(r"_x001D_+|\x1d+")


def split_field(val, max_parts=None):
    if val is None:
        return []
    parts = [p.strip() for p in SEP.split(str(val)) if p.strip()]
    parts = [p for p in parts if p and p != "x001D_"]
    return parts[:max_parts] if max_parts else parts


def clean_year(s):
    if not s:
        return ""
    return str(s).split("_x001D_")[0].strip().rstrip("?")


def load_existing(html):
    start = html.index("const DATA = ") + len("const DATA = ")
    end = html.index(";\n", start)
    data = json.loads(html[start:end])
    url_by_n = {d["n"]: d.get("url", "") for d in data}
    nfsa_by_n = {d["n"]: d.get("nfsa", "") for d in data}
    nfsa_by_cat = {
        str(d.get("cat", "")): (d.get("nfsa", ""), d.get("url", ""))
        for d in data
        if d.get("cat")
    }
    return data, url_by_n, nfsa_by_n, nfsa_by_cat, end


def import_excel(excel_path, url_by_n, nfsa_by_n, nfsa_by_cat):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = next(ws.iter_rows(max_row=1, values_only=True))
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        try:
            n = str(int(r["ID No."]))
        except (TypeError, ValueError):
            continue
        fn = (r.get("Performer - first name") or "").strip()
        sn = (r.get("Performer - second name or group") or "").strip()
        performer = f"{fn} {sn}".strip() if fn else sn
        releases = int(r.get("No.of releases") or 1)
        label = split_field(r.get("Label"), releases)
        catalogue = split_field(r.get("Catalogue ID"), releases)
        fmt = split_field(r.get("Format"), releases)
        year = split_field(r.get("Release date"), releases)
        album = (r.get("Album") or "").strip()
        genre = (r.get("Genre") or "").strip()
        if not performer and not album and not split_field(r.get("Label")):
            continue
        nfsa_num = r.get("NFSA Catalogue ID")
        cat = str(int(nfsa_num)) if nfsa_num not in (None, "") else ""
        url = url_by_n.get(n, "")
        nfsa = nfsa_by_n.get(n, "")
        if not url and cat and cat in nfsa_by_cat:
            nfsa, url = nfsa_by_cat[cat]
        rows.append(
            {
                "n": n,
                "performer": performer,
                "album": album,
                "label": label[0] if label else "",
                "catalogue": catalogue[0] if catalogue else "",
                "year": clean_year(year[0] if year else ""),
                "fmt": fmt[0] if fmt else "",
                "genre": genre,
                "cat": cat,
                "nfsa": nfsa,
                "url": url,
            }
        )
    wb.close()
    rows.sort(key=lambda d: int(d["n"]))
    return rows


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if not excel_path or not excel_path.exists():
        sys.exit(f"Usage: {sys.argv[0]} path/to/Matildas.xlsx")
    html = RECORDINGS.read_text()
    _, url_by_n, nfsa_by_n, nfsa_by_cat, _ = load_existing(html)
    rows = import_excel(excel_path, url_by_n, nfsa_by_n, nfsa_by_cat)
    data_js = "const DATA = " + json.dumps(rows, ensure_ascii=False) + ";"
    start = html.index("const DATA = ")
    end = html.index(";\n", start) + 1
    RECORDINGS.write_text(html[:start] + data_js + html[end:])
    print(f"Imported {len(rows)} recordings into {RECORDINGS}")


if __name__ == "__main__":
    main()
